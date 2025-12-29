from django.contrib.auth import authenticate
from django.db import IntegrityError
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from rest_framework import generics, permissions, status
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from .models import TeacherProfile, Attendance, Absence, Dropout, UnauthorizedPerson, ScanPhoto
from .serializers import (
    TeacherProfileSerializer,
    AttendanceSerializer,
    AbsenceSerializer,
    DropoutSerializer,
    UnauthorizedPersonSerializer,
    ScanPhotoSerializer
)
try:
    from openpyxl.cell.cell import MergedCell
except ImportError:
    MergedCell = type(None)  # fallback so isinstance won't fail
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.drawing.fill import GradientFillProperties, GradientStop
from datetime import datetime
from collections import defaultdict
from calendar import monthrange
from zoneinfo import ZoneInfo
import io
import json
import re

# ========================================
# TEACHER REGISTRATION (Public)
# ========================================
class RegisterView(generics.CreateAPIView):
    """Register a new teacher account"""
    queryset = TeacherProfile.objects.all()
    serializer_class = TeacherProfileSerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
            teacher_profile = serializer.save()
            token, _ = Token.objects.get_or_create(user=teacher_profile.user)
            return Response({
                "message": "Registration successful!",
                "token": token.key,
                "teacher": {
                    "id": teacher_profile.id,
                    "name": teacher_profile.user.first_name or teacher_profile.user.username,
                    "username": teacher_profile.user.username,
                    "section": teacher_profile.section,
                }
            }, status=status.HTTP_201_CREATED)
        except IntegrityError:
            return Response(
                {"error": "Username already exists."},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"Registration failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# ========================================
# TEACHER LOGIN (Public)
# ========================================
class LoginView(APIView):
    """Authenticate teacher and return token"""
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        username = request.data.get("username") or request.data.get("name")
        password = request.data.get("password")
        grade = request.data.get("grade")

        if not username or not password:
            return Response(
                {"error": "Username and password are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        user = authenticate(username=username, password=password)
        if user:
            try:
                teacher_profile = TeacherProfile.objects.get(user=user)
                if grade and grade.lower() not in teacher_profile.section.lower():
                    return Response(
                        {"error": "Grade does not match your assigned section"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                token, _ = Token.objects.get_or_create(user=user)
                return Response({
                    "token": token.key,
                    "teacher": {
                        "id": teacher_profile.id,
                        "name": user.first_name or username,
                        "username": username,
                        "section": teacher_profile.section,
                    }
                }, status=status.HTTP_200_OK)
            except TeacherProfile.DoesNotExist:
                return Response(
                    {"error": "Teacher profile not found"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return Response(
            {"error": "Invalid credentials"},
            status=status.HTTP_400_BAD_REQUEST
        )

# ========================================
# ATTENDANCE VIEWS
# ========================================
class AttendanceView(APIView):
    """List and create attendance records"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Get all attendance records with optional filters"""
        try:
            teacher_profile = TeacherProfile.objects.filter(user=request.user).first()
            if not teacher_profile:
                return Response(
                    {"error": "Teacher profile not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Apply filters
            date = request.query_params.get('date')
            student = request.query_params.get('student')
            status_filter = request.query_params.get('status')
            transaction_type = request.query_params.get('transaction_type')

            queryset = Attendance.objects.filter(teacher=teacher_profile)

            if date:
                queryset = queryset.filter(date=date)
            if student:
                queryset = queryset.filter(student_name__icontains=student)
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            if transaction_type:
                queryset = queryset.filter(transaction_type=transaction_type)

            attendances = queryset.order_by('-date', '-timestamp')
            serializer = AttendanceSerializer(attendances, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return Response(
                {"error": f"Error fetching attendance: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request):
        """Create a new attendance record"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            data = request.data.copy()
            qr_data = data.get('qr_data', '')

            # Parse QR code data if provided
            if qr_data:
                try:
                    qr_json = json.loads(qr_data)
                    data['student_lrn'] = qr_json.get('lrn', '')
                    if not data.get('student_name'):
                        data['student_name'] = qr_json.get('student', 'Unknown')

                    # ✅ EXTRACT GENDER FROM QR CODE
                    qr_gender = qr_json.get('gender', '').strip().upper()
                    if qr_gender:
                        # Convert F/M to Female/Male
                        if qr_gender == 'F' or qr_gender == 'FEMALE':
                            data['gender'] = 'Female'
                        elif qr_gender == 'M' or qr_gender == 'MALE':
                            data['gender'] = 'Male'

                    # ✅ EXTRACT GUARDIAN NAME FROM QR CODE
                    guardian_name = qr_json.get('name', '').strip()
                    guardian_role = qr_json.get('role', '').strip()
                    if guardian_name:
                        data['guardian_name'] = guardian_name

                except json.JSONDecodeError:
                    pass

            # Set default date if not provided
            if not data.get('date'):
                data['date'] = datetime.now().date()

            # Determine session based on Philippine Time if not provided
            if not data.get('session'):
                now = datetime.now()
                ph_time = now.astimezone(ZoneInfo('Asia/Manila'))
                data['session'] = 'AM' if ph_time.hour < 12 else 'PM'

            # ✅ NEW: Determine transaction type based on status
            status_value = data.get('status', 'Present')
            if status_value == 'Drop-off':
                data['transaction_type'] = 'drop-off'
            elif status_value == 'Pick-up':
                data['transaction_type'] = 'pick-up'
            else:
                data['transaction_type'] = 'attendance'

            serializer = AttendanceSerializer(data=data)
            if serializer.is_valid():
                attendance = serializer.save(teacher=teacher_profile)
                # send server-side push to parents if tokens available
                try:
                    from devices.expo import notify_parents_of_attendance
                    try:
                        notify_parents_of_attendance(attendance)
                    except Exception:
                        # avoid breaking the API response if push fails
                        pass
                except Exception:
                    pass
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([permissions.IsAuthenticated])
def attendance_detail(request, pk):
    """Retrieve, update, or delete a specific attendance record"""
    try:
        teacher_profile = TeacherProfile.objects.get(user=request.user)
        attendance = get_object_or_404(Attendance, pk=pk)

        if request.method == 'GET':
            serializer = AttendanceSerializer(attendance)
            return Response(serializer.data)

        elif request.method in ['PUT', 'PATCH']:
            partial = request.method == 'PATCH'
            
            # ✅ NEW: Update transaction_type if status is being changed
            data = request.data.copy()
            if 'status' in data:
                status_value = data['status']
                if status_value == 'Drop-off':
                    data['transaction_type'] = 'drop-off'
                elif status_value == 'Pick-up':
                    data['transaction_type'] = 'pick-up'
                else:
                    data['transaction_type'] = 'attendance'
            
            serializer = AttendanceSerializer(
                attendance,
                data=data,
                partial=partial
            )
            if serializer.is_valid():
                serializer.save(teacher=teacher_profile)
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            attendance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    except TeacherProfile.DoesNotExist:
        return Response(
            {"error": "Teacher profile not found"},
            status=status.HTTP_404_NOT_FOUND
        )
    except Attendance.DoesNotExist:
        return Response(
            {"error": "Attendance record not found"},
            status=status.HTTP_404_NOT_FOUND
        )

# ========================================
# PUBLIC ATTENDANCE LIST
# ========================================
class PublicAttendanceListView(generics.ListAPIView):
    """Public endpoint to view all attendance records (no authentication required)"""
    queryset = Attendance.objects.all().order_by('-timestamp')
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

# ========================================
# ABSENCE VIEWS
# ========================================
class AbsenceView(APIView):
    """List and create absence records"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Get all absence records for the authenticated teacher"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            absences = Absence.objects.filter(teacher=teacher_profile).order_by('-date')
            serializer = AbsenceSerializer(absences, many=True)
            return Response(serializer.data)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request):
        """Create a new absence record"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            serializer = AbsenceSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(teacher=teacher_profile)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )

@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([permissions.IsAuthenticated])
def absence_detail(request, pk):
    """Retrieve, update, or delete a specific absence record"""
    try:
        teacher_profile = TeacherProfile.objects.get(user=request.user)
        absence = get_object_or_404(Absence, pk=pk, teacher=teacher_profile)

        if request.method == 'GET':
            serializer = AbsenceSerializer(absence)
            return Response(serializer.data)

        elif request.method in ['PUT', 'PATCH']:
            partial = request.method == 'PATCH'
            serializer = AbsenceSerializer(
                absence,
                data=request.data,
                partial=partial
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            absence.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    except TeacherProfile.DoesNotExist:
        return Response(
            {"error": "Teacher profile not found"},
            status=status.HTTP_404_NOT_FOUND
        )
    except Absence.DoesNotExist:
        return Response(
            {"error": "Absence not found"},
            status=status.HTTP_404_NOT_FOUND
        )

# ========================================
# DROPOUT VIEWS
# ========================================
class DropoutView(APIView):
    """List and create dropout records"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Get all dropout records for the authenticated teacher"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            dropouts = Dropout.objects.filter(teacher=teacher_profile).order_by('-date')
            serializer = DropoutSerializer(dropouts, many=True)
            return Response(serializer.data)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request):
        """Create a new dropout record"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            serializer = DropoutSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(teacher=teacher_profile)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )

@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([permissions.IsAuthenticated])
def dropout_detail(request, pk):
    """Retrieve, update, or delete a specific dropout record"""
    try:
        teacher_profile = TeacherProfile.objects.get(user=request.user)
        dropout = get_object_or_404(Dropout, pk=pk, teacher=teacher_profile)

        if request.method == 'GET':
            serializer = DropoutSerializer(dropout)
            return Response(serializer.data)

        elif request.method in ['PUT', 'PATCH']:
            partial = request.method == 'PATCH'
            serializer = DropoutSerializer(
                dropout,
                data=request.data,
                partial=partial
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            dropout.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    except TeacherProfile.DoesNotExist:
        return Response(
            {"error": "Teacher profile not found"},
            status=status.HTTP_404_NOT_FOUND
        )
    except Dropout.DoesNotExist:
        return Response(
            {"error": "Dropout not found"},
            status=status.HTTP_404_NOT_FOUND
        )

# ========================================
# UNAUTHORIZED PERSON VIEWS
# ========================================
class UnauthorizedPersonView(APIView):
    """List and create unauthorized person records"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Get all unauthorized person records for the authenticated teacher"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            persons = UnauthorizedPerson.objects.filter(
                teacher=teacher_profile
            ).order_by('-timestamp')
            serializer = UnauthorizedPersonSerializer(persons, many=True)
            return Response(serializer.data)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request):
        """Create a new unauthorized person record"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            serializer = UnauthorizedPersonSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(teacher=teacher_profile)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )

@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([permissions.IsAuthenticated])
def unauthorized_person_detail(request, pk):
    """Retrieve, update, or delete a specific unauthorized person record"""
    try:
        teacher_profile = TeacherProfile.objects.get(user=request.user)
        person = get_object_or_404(
            UnauthorizedPerson,
            pk=pk,
            teacher=teacher_profile
        )

        if request.method == 'GET':
            serializer = UnauthorizedPersonSerializer(person)
            return Response(serializer.data)

        elif request.method in ['PUT', 'PATCH']:
            partial = request.method == 'PATCH'
            serializer = UnauthorizedPersonSerializer(
                person,
                data=request.data,
                partial=partial
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            person.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    except TeacherProfile.DoesNotExist:
        return Response(
            {"error": "Teacher profile not found"},
            status=status.HTTP_404_NOT_FOUND
        )
    except UnauthorizedPerson.DoesNotExist:
        return Response(
            {"error": "Person not found"},
            status=status.HTTP_404_NOT_FOUND
        )


# ========================================
# SF2 EXCEL REPORT GENERATION
# ========================================
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_sf2_excel(request):
    """
    Generate SF2 Excel report with attendance data for a specific month.
    Separates students by gender: Boys start at row 14, Girls start at row 36.
    ONLY WEEKDAYS (Monday-Friday) are included in the calendar.
    
    NAME FORMAT:
    - Column B (Row 14+): FULL NAME (Last Name, First Name Middle Name)
    
    Visual Legend:
    - AM Present (Morning): Green triangle ◤ - Top-left inside cell
    - PM Present (Afternoon): Green triangle ◢ - Bottom-right inside cell
    - Full Day Present (AM + PM): Solid green fill
    - Absent: Solid red fill
    
    Request Parameters:
    - template_file: Excel template file (multipart/form-data)
    - month: Optional, integer 1-12 (defaults to current month)
    - year: Optional, integer (defaults to current year)
    """
    try:
        teacher_profile = TeacherProfile.objects.get(user=request.user)
        template_file = request.FILES.get('template_file')
        
        if not template_file:
            return Response({"error": "Please upload an SF2 template file."}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = load_workbook(template_file)
        except Exception as e:
            return Response({"error": f"Failed to load Excel template: {str(e)}"}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            month = int(request.POST.get('month', datetime.now().month))
            year = int(request.POST.get('year', datetime.now().year))
        except ValueError:
            return Response({"error": "Invalid month or year parameter"}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if month < 1 or month > 12:
            return Response({"error": "Month must be between 1 and 12"}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        month_names = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", 
                      "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
        day_names_short = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        
        attendances = Attendance.objects.filter(
            teacher=teacher_profile,
            date__year=year,
            date__month=month
        ).order_by('date', 'timestamp')
        
        print(f"📊 Fetching attendance for: {month_names[month-1]} {year}")
        print(f"📝 Found {attendances.count()} attendance records")
        
        attendance_data = defaultdict(
            lambda: {'days': defaultdict(lambda: {'am': False, 'pm': False}), 'gender': None}
        )
        students_dict = {}
        
        for att in attendances:
            student_name = att.student_name
            day = att.date.day
            
            if student_name not in students_dict:
                students_dict[student_name] = getattr(att, 'gender', 'Male')
            
            if hasattr(att, 'session') and att.session:
                session = att.session.upper()
            elif att.timestamp:
                ph_time = att.timestamp.astimezone(ZoneInfo('Asia/Manila'))
                session = 'AM' if ph_time.hour < 12 else 'PM'
            else:
                session = 'AM'
            
            if att.status and att.status.lower() != 'absent':
                if session == 'AM':
                    attendance_data[student_name]['days'][day]['am'] = True
                elif session == 'PM':
                    attendance_data[student_name]['days'][day]['pm'] = True
            
            attendance_data[student_name]['gender'] = students_dict[student_name]
        
        boys = sorted([name for name, gender in students_dict.items() if gender.lower() == 'male'])
        girls = sorted([name for name, gender in students_dict.items() if gender.lower() == 'female'])
        
        print(f"👦 Boys: {len(boys)} students")
        print(f"👧 Girls: {len(girls)} students")
        
        now = datetime.now()
        current_day, current_month, current_year = now.day, now.month, now.year
        
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        
        # Optimized font size - 32pt with minimal shrinking needed
        triangle_font = Font(color="00B050", size=32, bold=True)
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # LOCKED TRIANGLE POSITIONS - Corners positioned precisely with shrink to fit enabled
        # AM triangle (◤) flush to top-left corner
        am_triangle_alignment = Alignment(
            horizontal='left', 
            vertical='top',
            indent=0,
            wrap_text=False,
            shrink_to_fit=True,  # Enable shrink to fit for scaling
            text_rotation=0
        )
        # PM triangle (◢) flush to bottom-right corner
        pm_triangle_alignment = Alignment(
            horizontal='right', 
            vertical='bottom',
            indent=0,
            wrap_text=False,
            shrink_to_fit=True,  # Enable shrink to fit for scaling
            text_rotation=0
        )
        
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
            print(f"📄 Processing sheet: {ws.title}")
        else:
            return Response({"error": "No sheets found in template"}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        date_row, day_row = 11, 12
        boys_start_row, girls_start_row = 14, 36
        name_column, first_day_column = 2, 4
        
        # Lock column widths for consistent triangle rendering
        # Set a uniform width for date columns to create square cells
        ATTENDANCE_COLUMN_WIDTH = 3.5  # Optimal width for triangle display
        
        def unmerge_and_write(ws, row, col, value, alignment=None):
            cell_coord = ws.cell(row=row, column=col).coordinate
            for merged_range in list(ws.merged_cells.ranges):
                if cell_coord in merged_range:
                    ws.unmerge_cells(str(merged_range))
                    print(f"  🔓 Unmerged {merged_range}")
                    break
            
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                del ws._cells[(row, col)]
                cell = ws.cell(row=row, column=col)
            
            cell.value = value
            if alignment:
                cell.alignment = alignment
            return cell
        
        days_in_month = monthrange(year, month)[1]
        day_columns = {}
        current_col = first_day_column
        
        print("\n📅 Filling weekday calendar headers (Mon-Fri only)...")
        for day in range(1, days_in_month + 1):
            current_date = date(year, month, day)
            weekday = current_date.weekday()
            
            if weekday < 5:
                day_columns[day] = current_col
                unmerge_and_write(ws, date_row, current_col, day, center_alignment)
                unmerge_and_write(ws, day_row, current_col, day_names_short[weekday], center_alignment)
                
                # Lock column width for consistent triangle positioning
                col_letter = ws.cell(row=1, column=current_col).column_letter
                ws.column_dimensions[col_letter].width = ATTENDANCE_COLUMN_WIDTH
                
                print(f"  Day {day:2d} ({current_date}) mapped to column {current_col} (width locked)")
                current_col += 1
            else:
                print(f"  Day {day:2d} ({current_date}) WEEKEND - skipped")
        
        print(f"✓ Filled {len(day_columns)} weekday columns with locked widths")
        
        # Lock row heights for consistent triangle display
        # Match row height to column width for square cells (perfect diagonal triangles)
        ROW_HEIGHT = 22  # Creates square cells optimized for 28pt triangles
        
        def is_merged_cell(ws, row, col):
            return isinstance(ws.cell(row=row, column=col), MergedCell)
        
        def fill_student_attendance(students_list, start_row):
            filled_count = 0
            for idx, name in enumerate(students_list):
                row_num = start_row + idx
                
                # Lock row height for consistent triangle positioning
                ws.row_dimensions[row_num].height = ROW_HEIGHT
                
                print(f"  Processing student: {name} at row {row_num} (height locked)")
                
                unmerge_and_write(ws, row_num, name_column, name, left_alignment)
                
                for day, col_idx in day_columns.items():
                    if year == current_year and month == current_month and day > current_day:
                        continue
                    
                    if is_merged_cell(ws, row_num, col_idx):
                        print(f"    ⏭️ Skipping merged cell at {row_num},{col_idx}")
                        continue
                    
                    cell = ws.cell(row=row_num, column=col_idx)
                    has_am = attendance_data[name]['days'][day]['am']
                    has_pm = attendance_data[name]['days'][day]['pm']
                    
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font()
                    cell.alignment = center_alignment
                    
                    if not has_am and not has_pm:
                        cell.fill = red_fill
                    elif has_am and has_pm:
                        cell.fill = green_fill
                    elif has_am and not has_pm:
                        cell.value = "◤"
                        cell.font = triangle_font
                        cell.alignment = am_triangle_alignment
                        cell.number_format = '@'
                        print(f"    ✓ AM triangle (◤) for day {day} - LOCKED POSITION")
                    elif has_pm and not has_am:
                        cell.value = "◢"
                        cell.font = triangle_font
                        cell.alignment = pm_triangle_alignment
                        cell.number_format = '@'
                        print(f"    ✓ PM triangle (◢) for day {day} - LOCKED POSITION")
                    
                    filled_count += 1
            
            return filled_count
        
        print(f"\n👦 Filling boys section starting at row {boys_start_row}")
        boys_filled = fill_student_attendance(boys, boys_start_row)
        print(f"✓ Filled {boys_filled} cells for boys with locked dimensions")
        
        print(f"\n👧 Filling girls section starting at row {girls_start_row}")
        girls_filled = fill_student_attendance(girls, girls_start_row)
        print(f"✓ Filled {girls_filled} cells for girls with locked dimensions")
        
        # Enable sheet protection to lock formatting (optional)
        # ws.protection.sheet = True
        # ws.protection.password = None  # No password, just lock structure
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"SF2_{month_names[month-1]}_{year}_{teacher_profile.section.replace(' ', '_')}.xlsx"
        
        print(f"\n✅ SF2 generated successfully: {filename}")
        print(f"📊 Total cells filled: {boys_filled + girls_filled}")
        print(f"🔒 All triangle positions LOCKED with uniform dimensions")
        
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except TeacherProfile.DoesNotExist:
        return Response({"error": "Teacher profile not found"}, 
                       status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        import traceback
        print("="*80)
        print("SF2 Generation Error:")
        print(traceback.format_exc())
        print("="*80)
        return Response({"error": f"Failed to generate SF2 Excel: {str(e)}"}, 
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# Add these corrected view classes at the end of your views.py file
# Replace the existing MarkUnscannedAbsentView, BulkMarkAbsentView, and AbsenceStatsView
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.decorators import api_view, permission_classes
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.utils import timezone
from datetime import date, datetime, timedelta
import logging

# Import your models
from parents.models import Student
from teacher.models import TeacherProfile, Attendance, Absence
from teacher.serializers import AbsenceSerializer

logger = logging.getLogger(__name__)


# ========================================
# MARK UNSCANNED ABSENT VIEW
# ========================================
class MarkUnscannedAbsentView(APIView):
    """
    Automatically mark all unscanned students as absent for a given date.
    Teacher is inferred from request.user, no teacher_id needed.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            # Parse date or use today
            target_date_str = request.data.get('date')
            if target_date_str:
                try:
                    target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response(
                        {"error": "Invalid date format. Use YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                target_date = date.today()

            # Get authenticated teacher profile
            try:
                teacher = TeacherProfile.objects.get(user=request.user)
            except TeacherProfile.DoesNotExist:
                return Response(
                    {"error": "Teacher profile not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            logger.info(f"Processing auto-absence for teacher {teacher.id} on {target_date}")

            # Get all students assigned to this teacher
            all_students = Student.objects.filter(teacher=teacher)
            total_students_count = all_students.count()
            if total_students_count == 0:
                return Response({
                    "message": "No students found in roster for this teacher",
                    "date": str(target_date),
                    "marked_count": 0,
                    "already_marked": 0,
                    "total_students": 0,
                    "students_marked": []
                })

            # Existing attendance records for this date
            existing_records = Attendance.objects.filter(date=target_date, teacher=teacher)
            scanned_lrns = set(existing_records.values_list('student_lrn', flat=True))

            # Students who haven't been scanned today
            unscanned_students = all_students.exclude(lrn__in=scanned_lrns)
            marked_students = []
            marked_count = 0

            with transaction.atomic():
                for student in unscanned_students:
                    record = Attendance.objects.create(
                        student_name=student.name,
                        student_lrn=student.lrn,
                        gender=student.gender or 'Male',
                        date=target_date,
                        status='Absent',
                        teacher=teacher,
                        session='AM',
                        transaction_type='attendance'
                    )
                    marked_students.append({
                        'lrn': student.lrn,
                        'name': student.name,
                        'gender': student.gender or '',
                        'reason': 'Not scanned - Auto-marked absent'
                    })
                    marked_count += 1

            already_marked = len(scanned_lrns)

            return Response({
                "message": f"Marked {marked_count} students as absent",
                "date": str(target_date),
                "marked_count": marked_count,
                "already_marked": already_marked,
                "total_students": total_students_count,
                "students_marked": marked_students
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception("Error in MarkUnscannedAbsentView")
            return Response(
                {"error": f"Failed to mark absences: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ========================================
# BULK MARK ABSENT VIEW
# ========================================
class BulkMarkAbsentView(APIView):
    """
    Bulk mark multiple dates for absence tracking.
    Teacher is inferred from request.user.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            start_date_str = request.data.get('start_date')
            end_date_str = request.data.get('end_date')

            if not start_date_str or not end_date_str:
                return Response(
                    {"error": "start_date and end_date are required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if start_date > end_date:
                return Response(
                    {"error": "start_date must be before end_date"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get authenticated teacher profile
            try:
                teacher = TeacherProfile.objects.get(user=request.user)
            except TeacherProfile.DoesNotExist:
                return Response(
                    {"error": "Teacher profile not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            current_date = start_date
            results = []

            while current_date <= end_date:
                # Use MarkUnscannedAbsentView logic for each date
                view = MarkUnscannedAbsentView()
                from rest_framework.request import Request
                from django.http import HttpRequest

                mock_request = HttpRequest()
                mock_request.user = request.user
                mock_request.data = {"date": current_date.strftime('%Y-%m-%d')}
                mock_request = Request(mock_request)
                mock_request._data = {"date": current_date.strftime('%Y-%m-%d')}

                response = view.post(mock_request)

                if response.status_code == 200:
                    results.append({
                        'date': str(current_date),
                        'marked': response.data.get('marked_count', 0)
                    })

                current_date += timedelta(days=1)

            total_marked = sum(r['marked'] for r in results)

            return Response({
                "message": f"Bulk marking complete. Marked {total_marked} total absences",
                "results": results,
                "total_marked": total_marked
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception("Error in BulkMarkAbsentView")
            return Response(
                {"error": f"Failed to bulk mark absences: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ========================================
# ABSENCE STATS VIEW
# ========================================
class AbsenceStatsView(APIView):
    """
    Get absence statistics for the authenticated teacher.
    Teacher is inferred from request.user.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            # Parse date
            target_date_str = request.query_params.get('date')
            if target_date_str:
                try:
                    target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
                except ValueError:
                    target_date = date.today()
            else:
                target_date = date.today()

            # Get authenticated teacher
            try:
                teacher = TeacherProfile.objects.get(user=request.user)
            except TeacherProfile.DoesNotExist:
                return Response(
                    {"error": "Teacher profile not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # All students for this teacher
            all_students = Student.objects.filter(teacher=teacher)
            total_students = all_students.count()

            # Attendance records for this date
            records_today = Attendance.objects.filter(date=target_date, teacher=teacher)
            present_count = records_today.filter(status='Present').count()
            absent_count = records_today.filter(status='Absent').count()
            scanned_lrns = set(records_today.values_list('student_lrn', flat=True))
            unscanned_count = total_students - len(scanned_lrns)

            scanned_percentage = round(
                (len(scanned_lrns) / total_students * 100) if total_students > 0 else 0,
                1
            )

            return Response({
                "date": str(target_date),
                "total_students": total_students,
                "present": present_count,
                "absent": absent_count,
                "unscanned": unscanned_count,
                "scanned_percentage": scanned_percentage
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception("Error in AbsenceStatsView")
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ========================================
# SCAN PHOTO VIEW
# ========================================
class ScanPhotoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Get all scan photos for the authenticated teacher"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            photos = ScanPhoto.objects.filter(
                teacher=teacher_profile
            ).order_by('-timestamp')
            serializer = ScanPhotoSerializer(photos, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request):
        """Save scan photo"""
        try:
            teacher_profile = TeacherProfile.objects.get(user=request.user)
            serializer = ScanPhotoSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(teacher=teacher_profile)
                return Response(
                    {"message": "Photo saved successfully"},
                    status=status.HTTP_201_CREATED
                )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TeacherProfile.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )
